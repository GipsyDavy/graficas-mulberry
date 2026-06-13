"""Script Vía A2 — Extrae hoja MATERIAL de PRECIOS PAPEL PROVEEDORES Formulas.xlsx
Genera: 3_materiales_plasticos_tintas.csv (nombre, categoria, unidad, precio_unidad, proveedor)
Compatible con Material.IMPORT_SPEC del wizard de importación.
"""
import openpyxl
import csv
import unicodedata
from pathlib import Path

EXCEL = Path(r"C:\Users\Gipsy Dávy\Desktop\excel\PRECIOS PAPEL PROVEEDORES Formulas.xlsx")
OUT   = Path(r"C:\Users\Gipsy Dávy\Desktop\excel\3_materiales_plasticos_tintas.csv")

PROVEEDORES_CONOCIDOS = {"CODIAL", "UNION PAPELERA", "UNION_PAPELERA", "UNIN PAPELERA"}


def normalizar(s):
    if s is None:
        return ""
    n = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return n.upper().strip()


def cell_val(ws, r, c):
    v = ws.cell(r, c).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def to_num(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    try:
        return round(float(str(v).replace(",", ".")), 4)
    except Exception:
        return ""


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb["MATERIAL"]
    rows = []

    # TINTAS (cols A-E, datos R4-R9)
    # A=MODELO, B=PRECIO, C=PRECIO_UD, D=REDONDEO, E=PROVEEDOR
    for r in range(4, 10):
        nombre = cell_val(ws, r, 1)
        if nombre is None:
            continue
        precio_ud = to_num(cell_val(ws, r, 3))
        proveedor = cell_val(ws, r, 5) or ""
        rows.append({
            "nombre": nombre,
            "categoria": "tintas",
            "unidad": "ud",
            "precio_unidad": precio_ud,
            "proveedor": proveedor,
        })

    # PLÁSTICOS (cols G-K)
    # R3 cabecera: G=MODELO, H=LONGITUD, I=PRECIO, J=PRECIO_UD, K=REDONDEO
    # Proveedor detectado cuando col G es nombre de empresa conocida
    proveedor_actual = ""
    for r in range(4, ws.max_row + 1):
        modelo = cell_val(ws, r, 7)
        if modelo is None:
            continue
        n = normalizar(modelo)
        if any(p in n for p in PROVEEDORES_CONOCIDOS):
            proveedor_actual = modelo
            continue
        if n in ("MODELO", "PLASTICO", "PLASTICOS"):
            continue
        precio_ud_raw = cell_val(ws, r, 10)
        precio_raw    = cell_val(ws, r, 9)
        if isinstance(precio_raw, str) and "NO" in precio_raw.upper():
            continue
        if isinstance(precio_ud_raw, str) and "NO" in precio_ud_raw.upper():
            precio_ud_raw = None
        precio_num = to_num(precio_ud_raw) if precio_ud_raw is not None else to_num(precio_raw)
        if precio_num == "":
            continue
        rows.append({
            "nombre": modelo,
            "categoria": "plastificado",
            "unidad": "m",
            "precio_unidad": precio_num,
            "proveedor": proveedor_actual,
        })

    # CONSUMIBLES (cols M-O: M=MODELO, N=CANTIDAD, O=PRECIO)
    for r in range(4, ws.max_row + 1):
        nombre = cell_val(ws, r, 13)
        if nombre is None:
            continue
        precio_raw = cell_val(ws, r, 15)
        if precio_raw is None:
            continue
        precio_num = to_num(precio_raw)
        if precio_num == "":
            continue
        cantidad = cell_val(ws, r, 14)
        if isinstance(cantidad, (int, float)):
            unidad_val = str(int(cantidad)) + " ud"
        else:
            unidad_val = str(cantidad) if cantidad else "ud"
        rows.append({
            "nombre": nombre,
            "categoria": "consumibles",
            "unidad": unidad_val,
            "precio_unidad": precio_num,
            "proveedor": "CODIAL",
        })

    with open(OUT, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["nombre", "categoria", "unidad", "precio_unidad", "proveedor"])
        writer.writeheader()
        writer.writerows(rows)

    print(f"Total: {len(rows)} filas -> {OUT}")
    for row in rows:
        cat = row["categoria"]
        nom = row["nombre"][:45]
        pre = row["precio_unidad"]
        prov = row["proveedor"]
        print(f"  [{cat}] {nom:<45} | precio={pre!s:>8} | proveedor={prov}")


if __name__ == "__main__":
    main()
